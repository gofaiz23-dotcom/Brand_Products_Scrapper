"""
Flow (matches Homelegance dealer site + your base sheet):

1. Load HomoleganceTool/.env: LOGIN_URL, USERNAME, PASSWORD (override=True so .env
   wins over Windows USERNAME).
2. Open LOGIN_URL once, fill username/password, sign in (single session for the run).
3. Read base xlsx (setting.get_base_sheet_path()); use column Sub-SKU.
4. Skip rows where Sub-SKU contains a comma (multi-SKU cells).
5. For each single Sub-SKU: search in the header, open the exact product card, scrape
   the PDP (#product_name, price, stock, description, images from #product_img / bpic_*,
   Weights & Dimensions / Product Details / Packaging from .model-info-box — not the
   "See more from … Collection" carousel).
6. New workbook each run under sheets/Homelagance/scrppedSheets: headers
   first, then only successfully scraped rows are appended and saved (live — open the
   file while running to watch). Skipped / failed SKUs are not written to this file.
"""

from __future__ import annotations

import logging
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any
from urllib.parse import urljoin, urlparse

import openpyxl
from openpyxl import Workbook
from dotenv import load_dotenv
from openpyxl.worksheet.worksheet import Worksheet
from playwright.sync_api import Page, sync_playwright

import setting

log = logging.getLogger(__name__)


def _esc(s: str) -> str:
    try:
        from rich.markup import escape

        return escape(str(s))
    except ImportError:
        return str(s).replace("[", "\\[")

# Appended after base-sheet columns (same order as scrape_product_page output).
SCRAPED_COLUMN_KEYS: list[str] = [
    "scrape_status",
    "scrape_error",
    "Title",
    "Brand real price",
    "Inventory",
    "Product Description",
    "main image",
    "gallery image",
    "Weights & Dimensions",
    "Product Details",
    "Packaging",
    "Product page URL",
]


def _load_env() -> None:
    env_path = setting.HOMETOOL_DIR / ".env"
    if env_path.is_file():
        # So HomoleganceTool/.env wins over Windows' built-in USERNAME, etc.
        load_dotenv(env_path, override=True)


def _origin_from_login_url(login_url: str) -> str:
    p = urlparse(login_url.strip())
    if p.scheme and p.netloc:
        return f"{p.scheme}://{p.netloc}"
    return setting.SITE_ORIGIN.rstrip("/")


def _abs_url(origin: str, href: str | None) -> str | None:
    if not href:
        return None
    h = href.strip()
    if not h:
        return None
    return urljoin(origin + "/", h.lstrip("/"))


def _norm_sku(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip()).casefold()


def is_single_sub_sku(cell_value: Any) -> bool:
    if cell_value is None:
        return False
    text = str(cell_value).strip()
    if not text:
        return False
    if "," in text:
        return False
    return True


def single_sub_sku_value(cell_value: Any) -> str:
    return re.sub(r"\s+", " ", str(cell_value).strip())


def count_single_sub_sku_rows(data_rows: list[list[Any]], sub_col: int) -> int:
    n = 0
    for row in data_rows:
        cell = row[sub_col] if sub_col < len(row) else None
        if is_single_sub_sku(cell):
            n += 1
    return n


def read_sheet_rows(path: Path) -> tuple[list[str], list[list[Any]]]:
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))
    finally:
        wb.close()
    if not rows:
        return [], []
    header = [str(c).strip() if c is not None else "" for c in rows[0]]
    data = [list(r) for r in rows[1:]]
    return header, data


def ensure_output_headers(ws: Worksheet, start_col: int) -> None:
    for i, key in enumerate(SCRAPED_COLUMN_KEYS):
        ws.cell(row=1, column=start_col + i, value=key)


SCRAPE_COL_OFFSET: dict[str, int] = {
    k: i for i, k in enumerate(SCRAPED_COLUMN_KEYS)
}


def new_output_path() -> Path:
    stamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    return setting.LIVE_SUCCESS_XLSX_DIR / f"Homelagance-{stamp}-scrapped-products.xlsx"


def write_scrape_cells(
    ws: Worksheet, excel_row: int, n_base: int, data: dict[str, str]
) -> None:
    start = n_base + 1
    for key in SCRAPED_COLUMN_KEYS:
        ws.cell(
            row=excel_row,
            column=start + SCRAPE_COL_OFFSET[key],
            value=data.get(key, ""),
        )


def init_success_only_workbook(
    out_path: Path, base_header: list[str]
) -> tuple[Workbook, Worksheet, int]:
    """Headers only; data rows are appended only when a scrape succeeds."""
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"

    n_base = len(base_header)
    for j, h in enumerate(base_header, start=1):
        ws.cell(row=1, column=j, value=h)
    ensure_output_headers(ws, n_base + 1)
    wb.save(out_path)
    return wb, ws, n_base


def append_success_row(
    ws: Worksheet,
    excel_row: int,
    n_base: int,
    base_row: list[Any],
    scrape_data: dict[str, str],
) -> None:
    for j in range(n_base):
        val = base_row[j] if j < len(base_row) else None
        ws.cell(row=excel_row, column=j + 1, value=val)
    merged = {k: "" for k in SCRAPED_COLUMN_KEYS}
    merged.update(scrape_data)
    merged["scrape_status"] = "ok"
    merged["scrape_error"] = ""
    write_scrape_cells(ws, excel_row, n_base, merged)


def login_once(
    page: Page, login_url: str, username: str, password: str
) -> None:
    """Log in a single time for this browser session. Do not call again per SKU."""
    log.info(
        "[bold blue]Login[/] → opening [underline cyan]%s[/underline cyan]",
        _esc(login_url),
    )
    page.goto(login_url, wait_until="domcontentloaded")
    page.wait_for_selector("#username", timeout=setting.DEFAULT_TIMEOUT_MS)
    log.info("[blue]Login[/] → [dim]filling username / password (one session for all SKUs)[/]")
    page.fill("#username", username)
    page.fill("#password", password)
    page.click("#submitBtn")
    page.wait_for_selector("#web_search_model", timeout=setting.DEFAULT_TIMEOUT_MS)
    log.info("[bold green]Login OK[/] — [green]header search box ready[/]")


def search_sku(page: Page, sku: str) -> None:
    """Run site search while already logged in (reuses session; no login)."""
    search = page.locator("#web_search_model")
    search.click()
    search.fill("")
    search.fill(sku)
    page.locator("#searchModelButton").click()
    page.wait_for_selector(
        "div.pro-wrapper.thumb-item-box",
        timeout=setting.DEFAULT_TIMEOUT_MS,
    )


def click_exact_product_card(page: Page, sku: str) -> bool:
    target = _norm_sku(sku)
    cards = page.locator(
        "div.pro-wrapper.thumb-item-box.quick-view-box:not(.pro-wrapper-collection)"
    )
    n = cards.count()
    for i in range(n):
        card = cards.nth(i)
        name_el = card.locator('p[id^="modelName"]').first
        if not name_el.count():
            name_el = card.locator('input[id^="bgitem_name_"]').first
        if not name_el.count():
            continue
        raw = name_el.get_attribute("value")
        if raw is None:
            raw = name_el.inner_text()
        if raw is None:
            continue
        if _norm_sku(raw) != target:
            continue
        link = card.locator("a.flex-fill").first
        if link.count():
            with page.expect_navigation(
                wait_until="domcontentloaded",
                timeout=setting.DEFAULT_TIMEOUT_MS,
            ):
                link.click()
            return True
    return False


def _text_clean(s: str) -> str:
    return re.sub(r"\s+", " ", s.strip())


def _format_brand_price(price_raw: str) -> str:
    s = (price_raw or "").strip()
    if not s:
        return ""
    try:
        return f"$ {float(s):.2f}"
    except ValueError:
        return f"$ {s}" if not s.lstrip().startswith("$") else s


def _norm_list_lines(page: Page, heading_fragment: str) -> str:
    """One list item per line (matches copy-paste style from the PDP)."""
    box = page.locator(
        "xpath=//p[contains(@class,'norm-name')]"
        f"[contains(normalize-space(.), '{heading_fragment}')]"
        "/following-sibling::ol[1]"
    ).first
    if not box.count():
        return ""
    items = box.locator("li")
    parts: list[str] = []
    for k in range(items.count()):
        line = _text_clean(items.nth(k).inner_text())
        if line:
            parts.append(line)
    return "\n".join(parts)


def _packaging_block(page: Page, origin: str) -> str:
    box = page.locator(
        "xpath=//p[contains(@class,'norm-name')]"
        "[contains(normalize-space(.), 'Packaging')]"
        "/following-sibling::ol[1]"
    ).first
    if not box.count():
        return ""
    items = box.locator("li")
    lines: list[str] = []
    for k in range(items.count()):
        li = items.nth(k)
        t = _text_clean(li.inner_text())
        pdf = li.locator("a[href*='viewPdf']").first
        if pdf.count():
            href = pdf.get_attribute("href")
            au = _abs_url(origin, href) or (href or "").strip()
            if au and au not in t:
                t = f"{t} {au}".strip() if t else au
        if t:
            lines.append(t)
    return "\n".join(lines)


def scrape_product_page(page: Page, origin: str) -> dict[str, str]:
    """Fields align with #hidden-vals + PDP sections (not related / collection carousels)."""
    page.wait_for_selector(
        "#hidden-vals", state="attached", timeout=setting.DEFAULT_TIMEOUT_MS
    )

    def gv(sel: str) -> str:
        el = page.locator(sel).first
        if el.count():
            v = el.get_attribute("value")
            return (v or "").strip()
        return ""

    product_name = gv("#product_name")
    product_desc_short = gv("#product_description")
    title = _text_clean(f"{product_name} {product_desc_short}".strip())

    price_raw = gv("#fn_price")
    if not price_raw:
        vis = page.locator(".model_price_nomal").first
        if vis.count():
            price_raw = _text_clean(vis.inner_text())
    brand_price = _format_brand_price(price_raw)

    santa_fe = gv("#availabilty")

    remote = ""
    remote_blk = page.locator("div.mt-2.pl-1").filter(
        has_text=re.compile(r"Remote\s+Warehouse\s+Stock", re.I)
    )
    if remote_blk.count():
        m = re.search(
            r"Remote\s+Warehouse\s+Stock\s*:?\s*(\d+)",
            remote_blk.first.inner_text(),
            re.I,
        )
        if m:
            remote = m.group(1)

    inventory = ""
    if santa_fe or remote:
        inventory = (
            f"Santa Fe Springs Warehouse Stock: {santa_fe or '0'} , "
            f"Remote Warehouse Stock : {remote or '0'}"
        )

    long_desc = ""
    ld = page.locator(".collapse-description.overflow-auto").first
    if ld.count():
        long_desc = _text_clean(ld.inner_text())

    weights = _norm_list_lines(page, "Weights")
    product_details = _norm_list_lines(page, "Product Details")
    packaging = _packaging_block(page, origin)

    main_src = ""
    img = page.locator("#product_img").first
    if img.count():
        main_src = img.get_attribute("src") or ""
    main_abs = _abs_url(origin, main_src) or main_src

    gallery: list[str] = []
    for inp in page.locator('#pc_show_image_div input[id^="bpic_"]').all():
        v = inp.get_attribute("value")
        if v:
            u = _abs_url(origin, v)
            if u:
                gallery.append(u)
    if not gallery:
        for inp in page.locator('#mb_show_image_div a.cloud-zoom-gallery').all():
            href = inp.get_attribute("href")
            u = _abs_url(origin, href)
            if u:
                gallery.append(u)

    dedup: list[str] = []
    seen: set[str] = set()
    for u in gallery:
        if u not in seen:
            seen.add(u)
            dedup.append(u)

    rest = [u for u in dedup if u != main_abs]

    return {
        "Title": title,
        "Brand real price": brand_price,
        "Inventory": inventory,
        "Product Description": long_desc,
        "main image": main_abs,
        "gallery image": ",".join(rest),
        "Weights & Dimensions": weights,
        "Product Details": product_details,
        "Packaging": packaging,
        "Product page URL": page.url,
    }


def run() -> Path:
    _load_env()
    login_url = os.environ.get("LOGIN_URL", "").strip()
    username = os.environ.get("USERNAME", "").strip()
    password = os.environ.get("PASSWORD", "").strip()
    if not login_url or not username or not password:
        raise SystemExit(
            "Set LOGIN_URL, USERNAME, and PASSWORD in HomoleganceTool/.env"
        )

    origin = _origin_from_login_url(login_url)

    base_path = setting.get_base_sheet_path()
    if not base_path.is_file():
        raise SystemExit(
            "Base Excel not found.\n"
            f"  Folder: {setting.BASE_SHEETS_DIR}\n"
            "  Put the file there as one of:\n"
            "    - Homelegance-All-CategoriesSkus-InSingleSheet (1).xlsx\n"
            "    - Homelegance-All-CategoriesSkus-InSingleSheet.xlsx\n"
            "    - or any other .xlsx (newest is used if several exist).\n"
            f"  Looked for: {base_path}"
        )

    base_header, data_rows = read_sheet_rows(base_path)
    if setting.SUB_SKU_HEADER not in base_header:
        raise SystemExit(f"Missing column {setting.SUB_SKU_HEADER!r} in base sheet")
    sub_col = base_header.index(setting.SUB_SKU_HEADER)

    total_rows = len(data_rows)
    single_sub_sku_total = count_single_sub_sku_rows(data_rows, sub_col)
    max_cap = setting.MAX_SKUS
    jobs_planned = (
        min(single_sub_sku_total, max_cap) if max_cap is not None else single_sub_sku_total
    )

    log.info("[bold white on blue] ═══ Homelegance single Sub-SKU scraper ═══ [/]")
    log.info("[green]Base sheet[/] → [cyan]%s[/cyan]", base_path)
    log.info("[magenta]Rows[/] in sheet (data): [bold]%d[/bold]", total_rows)
    log.info(
        "[yellow]Single Sub-SKU[/] rows (no comma): [bold cyan]%d[/] / [dim]%d[/] data rows",
        single_sub_sku_total,
        total_rows,
    )
    if max_cap is not None:
        log.info(
            "[yellow]MAX_SKUS[/]=[bold]%d[/] → will attempt [bold green]%d[/] scrapes",
            max_cap,
            jobs_planned,
        )
    else:
        log.info(
            "[yellow]MAX_SKUS[/]=[dim]unlimited[/] → [bold green]%d[/] scrapes planned",
            jobs_planned,
        )
    log.info("[blue]Site[/] → [underline cyan]%s[/underline cyan]", _esc(origin))

    out_path = new_output_path()
    log.info("[green]Live xlsx[/] (success rows only) → [cyan]%s[/cyan]", out_path)
    wb, ws, n_base = init_success_only_workbook(out_path, base_header)
    log.info("[dim]Workbook ready[/] [bold](headers only)[/] — rows append + save after each [green]OK[/]")
    next_data_row = 2

    single_sku_attempts = 0

    with sync_playwright() as p:
        browser = p.chromium.launch(
            headless=setting.HEADLESS,
            slow_mo=setting.SLOW_MO_MS,
        )
        context = browser.new_context(
            viewport={"width": 1400, "height": 900},
            ignore_https_errors=True,
        )
        page = context.new_page()
        page.set_default_timeout(setting.DEFAULT_TIMEOUT_MS)

        login_once(page, login_url, username, password)

        for row in data_rows:
            sku_cell = row[sub_col] if sub_col < len(row) else None

            if not is_single_sub_sku(sku_cell):
                continue

            sku = single_sub_sku_value(sku_cell)
            if setting.MAX_SKUS is not None and single_sku_attempts >= setting.MAX_SKUS:
                log.info(
                    "[yellow]Stopped[/] — reached [bold]MAX_SKUS=%d[/] (remaining singles skipped)",
                    setting.MAX_SKUS,
                )
                break

            single_sku_attempts += 1
            log.info(
                "[bold magenta]Single Sub-SKU[/] [bold cyan]%d[/]/[bold white]%d[/] "
                "| Sub-SKU [bold yellow]%s[/]",
                single_sku_attempts,
                jobs_planned,
                _esc(sku),
            )
            try:
                log.info("[blue]Search[/] → typing Sub-SKU into header → [dim]submit[/]")
                search_sku(page, sku)
                log.info("[blue]Search[/] → [green]results grid[/] — matching exact card…")
                if click_exact_product_card(page, sku):
                    log.info(
                        "[bold green]PDP[/] → scraping [dim]title, price, stock, images, specs…[/]"
                    )
                    data = scrape_product_page(page, origin)
                    log.info(
                        "[green]Scraped[/] Title=[white]%s[/] | price=[yellow]%s[/] | "
                        "main image [dim]%d chars[/]",
                        _esc((data.get("Title") or "")[:80]),
                        _esc(str(data.get("Brand real price") or "")),
                        len(data.get("main image") or ""),
                    )
                    append_success_row(ws, next_data_row, n_base, row, data)
                    log.info(
                        "[bold green]XLSX[/] row [cyan]%d[/] saved → [magenta]%s[/]",
                        next_data_row,
                        out_path.name,
                    )
                    next_data_row += 1
                    wb.save(out_path)
                else:
                    log.warning(
                        "[bold red]FAIL[/] — no exact card for Sub-SKU [yellow]%s[/] "
                        "[dim](not written to xlsx)[/]",
                        _esc(sku),
                    )
            except Exception as e:  # noqa: BLE001
                log.warning(
                    "[bold red]ERROR[/] — Sub-SKU [yellow]%s[/]: [red]%s[/] "
                    "[dim](not written to xlsx)[/]",
                    _esc(sku),
                    _esc(str(e)),
                )
                try:
                    page.goto(origin + "/", wait_until="domcontentloaded")
                except Exception:
                    pass

            time.sleep(setting.DELAY_BETWEEN_SKUS_SEC)

        context.close()
        browser.close()

    successes = next_data_row - 2
    log.info(
        "[bold white on green] ═══ Finished ═══ [/] "
        "[green]success rows[/]=[bold]%d[/] | attempts [cyan]%d[/]/[white]%d[/] | file [magenta]%s[/]",
        successes,
        single_sku_attempts,
        jobs_planned,
        out_path,
    )
    return out_path


if __name__ == "__main__":
    from log_theme import setup_colored_logging

    setup_colored_logging()
    out = run()
    log.info("[bold green]Run complete.[/] Output: [cyan]%s[/cyan]", out)
