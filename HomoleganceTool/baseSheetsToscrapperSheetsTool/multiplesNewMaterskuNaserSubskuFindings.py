"""
After singleSUB_SKUscraper.run() writes the singles xlsx:

1. Re-read base sheet; rows where Sub-SKU contains a comma (multi).
2. If every Sub-SKU in that cell exists in the singles output (ok rows) → no browser:
   write one row to Homelagance-{stamp}-multiple-subskus-from-singles.xlsx with narrow
   columns (MULTI_NARROW_BASE_HEADERS + attributes JSON from singles only).
3. If any Sub-SKU is missing from singles → search NEW/Master SKU (strip HML-; 3+ segments
   drop last) on the site, scrape PDP + packaging table; write to
   Homelagance-{stamp}-multiple-subskus-sheets.xlsx (full base + scrape columns).
4. Web scrape ok: attributes = JSON array in Sub-SKU sheet order; each item is the full packaging
   row when Model matches that Sub-SKU (simple normalized match), else singles merge, else {model}.
   No table / failed scrape: same merge logic with an empty table.
"""

from __future__ import annotations

import json
import logging
import os
import re
import time
from datetime import datetime
from pathlib import Path
from typing import Any

import openpyxl
from playwright.sync_api import Page, sync_playwright

import setting
from singleSUB_SKUscraper import (
    SCRAPED_COLUMN_KEYS,
    _esc,
    _load_env,
    _norm_sku,
    _origin_from_login_url,
    _text_clean,
    append_scrape_row,
    click_exact_product_card,
    init_success_only_workbook,
    is_single_sub_sku,
    login_once,
    read_sheet_rows,
    scrape_product_page,
    search_sku,
)

log = logging.getLogger(__name__)

_SCRAPE_TO_JSON_EXTRA: dict[str, str] = {
    "Title": "title",
    "Brand real price": "brand_real_price",
    "Inventory": "inventory",
    "Product Description": "product_description",
    "main image": "main_image",
    "gallery image": "gallery_image",
    "Weights & Dimensions": "weights_dimensions",
    "Product Details": "product_details",
    "Packaging": "packaging",
    "Product page URL": "product_page_url",
}


def _norm_sku_match(s: str) -> str:
    return _norm_sku(str(s).strip().rstrip("*"))


def normalize_master_sku_for_search(raw: Any) -> str:
    s = str(raw or "").strip()
    if not s:
        return ""
    s = re.sub(r"^HML\s*-\s*", "", s, flags=re.IGNORECASE).strip()
    parts = s.split("-")
    if len(parts) >= 3:
        s = "-".join(parts[:-1])
    return s.strip()


def parse_multi_sub_skus(cell_value: Any) -> list[str]:
    if cell_value is None:
        return []
    text = str(cell_value).strip()
    if not text:
        return []
    return [
        re.sub(r"\s+", " ", p.strip())
        for p in text.split(",")
        if p.strip()
    ]


def count_multi_sub_sku_rows(data_rows: list[list[Any]], sub_col: int) -> int:
    n = 0
    for row in data_rows:
        cell = row[sub_col] if sub_col < len(row) else None
        if cell is None or str(cell).strip() == "":
            continue
        if is_single_sub_sku(cell):
            continue
        n += 1
    return n


def _parse_num(s: str) -> float | int | str:
    t = _text_clean(s)
    if not t:
        return ""
    t_price = re.sub(r"[^\d.\-]", "", t)
    if not t_price:
        return t
    try:
        f = float(t_price)
        if f == int(f):
            return int(f)
        return f
    except ValueError:
        return t


def scrape_packaging_el_table(page: Page) -> list[dict[str, Any]]:
    """Read the PDP packaging Element table (Model, Box, …).

    The grid is the ``div.row.mt-4`` immediately **after** ``.model-info-box`` (not other
    ``.row.mt-4`` carousels). Read **only** ``.el-table > .el-table__body-wrapper`` rows —
    not ``.el-table__fixed-body-wrapper``, where Box/price cells are ``is-hidden`` and
    often scrape empty. Use ``textContent`` for cells (includes hidden model column).
    """

    def _cell_text(td_loc: Any) -> str:
        inner = td_loc.locator(".cell").first
        if inner.count():
            raw = inner.text_content()
        else:
            raw = td_loc.text_content()
        return _text_clean(raw or "")

    def _main_body_rows(el_table: Any) -> Any:
        """Direct scroll body only (sibling of .el-table__fixed), not the fixed clone."""
        return el_table.locator(":scope > .el-table__body-wrapper tbody tr.el-table__row")

    row_after_specs = page.locator(".model-info-box ~ div.row.mt-4").filter(
        has=page.locator("th", has_text=re.compile(r"Model", re.I))
    )
    row_any_mt4 = page.locator("div.row.mt-4").filter(
        has=page.locator("th", has_text=re.compile(r"Model", re.I))
    )

    try:
        row_after_specs.first.scroll_into_view_if_needed(timeout=5_000)
    except Exception:
        try:
            row_any_mt4.first.scroll_into_view_if_needed(timeout=5_000)
        except Exception:
            pass

    try:
        page.wait_for_function(
            """() => {
              const table = document.querySelector(
                '.model-info-box ~ div.row.mt-4 div.el-table'
              ) || document.querySelector('div.row.mt-4 div.el-table');
              if (!table) return false;
              const wrap = table.querySelector(':scope > .el-table__body-wrapper tbody');
              if (!wrap) return false;
              const tr = wrap.querySelector('tr.el-table__row');
              if (!tr) return false;
              const tds = tr.querySelectorAll('td');
              if (tds.length < 8) return false;
              const box = (tds[1].textContent || '').trim();
              return box.length > 0;
            }""",
            timeout=20_000,
        )
    except Exception:
        try:
            page.wait_for_selector(
                ".model-info-box ~ div.row.mt-4 div.el-table > .el-table__body-wrapper "
                "tbody tr.el-table__row",
                state="attached",
                timeout=12_000,
            )
        except Exception:
            try:
                page.wait_for_selector(
                    "div.row.mt-4 div.el-table > .el-table__body-wrapper "
                    "tbody tr.el-table__row",
                    state="attached",
                    timeout=8_000,
                )
            except Exception:
                pass

    el_near = row_after_specs.locator("div.el-table").first
    el_wide = row_any_mt4.locator("div.el-table").first

    row_locator_candidates: list[Any] = [
        _main_body_rows(el_near),
        _main_body_rows(el_wide),
        page.locator(
            ".model-info-box ~ div.row.mt-4 div.el-table > .el-table__body-wrapper "
            "tbody tr.el-table__row"
        ),
        page.locator(
            "div.row.mt-4 div.el-table > .el-table__body-wrapper tbody tr.el-table__row"
        ),
        page.locator("div.el-table > .el-table__body-wrapper tbody tr.el-table__row"),
        row_any_mt4.locator(".el-table__fixed-body-wrapper tbody tr.el-table__row"),
        page.locator("table.el-table__body tbody tr.el-table__row"),
    ]

    def _pick_best(candidates: list[Any]) -> tuple[Any, int]:
        best_r: Any = candidates[-1]
        best_score = -1
        best_fill = -1
        for rloc in candidates:
            try:
                cnt = rloc.count()
            except Exception:
                continue
            score = 0
            fill0 = 0
            for ri in range(min(cnt, 50)):
                tr = rloc.nth(ri)
                if tr.locator("td").count() < 8:
                    continue
                score += 1
                if ri == 0:
                    for ci in range(1, 8):
                        if _cell_text(tr.locator("td").nth(ci)):
                            fill0 += 1
            if score > best_score or (
                score == best_score and score > 0 and fill0 > best_fill
            ):
                best_score = score
                best_fill = fill0
                best_r = rloc
        return best_r, best_score

    rows, best_score = _pick_best(row_locator_candidates)
    if best_score <= 0:
        rows = page.locator("tbody tr.el-table__row")

    out: list[dict[str, Any]] = []
    seen_row_models: set[str] = set()
    n = rows.count()
    for i in range(n):
        tr = rows.nth(i)
        tds = tr.locator("td")
        if tds.count() < 8:
            continue
        model = _cell_text(tds.nth(0))
        if not model:
            continue
        mk = _norm_sku_match(model)
        if mk in seen_row_models:
            continue
        seen_row_models.add(mk)

        box_raw = _cell_text(tds.nth(1))
        desc = _cell_text(tds.nth(2))
        cu = _cell_text(tds.nth(3))
        gw = _cell_text(tds.nth(4))
        nw = _cell_text(tds.nth(5))
        pkg = _cell_text(tds.nth(6))
        unit = _cell_text(tds.nth(7))

        box_v: Any = _parse_num(box_raw)
        if isinstance(box_v, str) and box_v != "":
            try:
                box_v = int(float(box_v))
            except ValueError:
                pass

        out.append(
            {
                "model": model,
                "box": box_v,
                "description": desc,
                "cuFt": _parse_num(cu),
                "grossWeight": _parse_num(gw),
                "netWeight": _parse_num(nw),
                "packageDimensions": pkg,
                "unitPrice": _parse_num(unit),
            }
        )
    return out


def _extras_from_singles_scrape(scrape: dict[str, str]) -> dict[str, str]:
    o: dict[str, str] = {}
    for src, dst in _SCRAPE_TO_JSON_EXTRA.items():
        v = (scrape.get(src) or "").strip()
        if v:
            o[dst] = v
    return o


def _attributes_ordered_with_sources(
    sub_skus: list[str],
    table_rows: list[dict[str, Any]],
    singles_by_sku: dict[str, dict[str, str]],
) -> tuple[list[dict[str, Any]], list[str]]:
    """Build ordered attribute dicts plus one log line per Sub-SKU (source: table / singles / model-only)."""
    by_model: dict[str, dict[str, Any]] = {}
    for tr in table_rows:
        m = str(tr.get("model") or tr.get("sub-sku") or "").strip()
        if not m:
            continue
        row_clean = {k: v for k, v in tr.items() if k != "sub-sku"}
        row_clean["model"] = m
        key = _norm_sku_match(m)
        by_model[key] = dict(row_clean)

    ordered: list[dict[str, Any]] = []
    log_lines: list[str] = []
    for sub in sub_skus:
        km = _norm_sku_match(sub)
        row_hit = by_model.get(km)
        if row_hit is not None:
            ordered.append(dict(row_hit))
            log_lines.append(f"{sub} → packaging table (model match)")
        elif km in singles_by_sku:
            cell = {"model": sub}
            cell.update(_extras_from_singles_scrape(singles_by_sku[km]))
            ordered.append(cell)
            log_lines.append(f"{sub} → singles file merge (no table row)")
        else:
            ordered.append({"model": sub})
            log_lines.append(f"{sub} → model only (no table match, not in singles)")
    return ordered, log_lines


def build_attributes_json(
    sub_skus: list[str],
    table_rows: list[dict[str, Any]],
    singles_by_sku: dict[str, dict[str, str]],
) -> str:
    """One JSON object per Sub-SKU (sheet order).

    Simple rule: normalize sheet Sub-SKU and table Model the same way (``_norm_sku_match``).
    If they match → full packaging row from HTML (model, box, description, …).
    Else if that Sub-SKU exists in singles output → merge singles extras under ``model``.
    Else → ``{"model": "<sub>"}`` only.
    """
    ordered, _ = _attributes_ordered_with_sources(sub_skus, table_rows, singles_by_sku)
    return json.dumps(ordered, ensure_ascii=False)


def log_attributes_breakdown(
    sub_skus: list[str],
    table_rows: list[dict[str, Any]],
    singles_by_sku: dict[str, dict[str, str]],
    *,
    mode: str,
) -> None:
    """Log how each Sub-SKU was resolved for the ``attributes`` column."""
    if not sub_skus:
        return
    _, lines = _attributes_ordered_with_sources(sub_skus, table_rows, singles_by_sku)
    log.info(
        "[cyan]attributes[/] [%s] %d Sub-SKU(s) — per-item source:",
        mode,
        len(sub_skus),
    )
    for ln in lines:
        log.info("  [dim]•[/] %s", _esc(ln))
    if table_rows:
        models = [str(r.get("model") or "").strip() for r in table_rows if r.get("model")]
        log.info(
            "  [dim]packaging table on page:[/] [white]%d[/] row(s) — models [dim]%s[/]",
            len(table_rows),
            _esc(", ".join(models) if models else "(none)"),
        )


def _read_header_row(ws: Any) -> list[str]:
    headers: list[str] = []
    for j in range(1, ws.max_column + 1):
        v = ws.cell(row=1, column=j).value
        headers.append(str(v).strip() if v is not None else "")
    return headers


def _n_base_from_headers(headers: list[str]) -> int:
    if "scrape_status" not in headers:
        raise SystemExit(
            "Live xlsx missing column 'scrape_status'. Run singleSUB_SKUscraper first."
        )
    return headers.index("scrape_status")


def load_singles_by_subsku(path: Path) -> tuple[dict[str, dict[str, str]], list[str], int]:
    wb = openpyxl.load_workbook(path, data_only=True)
    try:
        ws = wb.active
        headers = _read_header_row(ws)
        n_base = _n_base_from_headers(headers)
        if setting.SUB_SKU_HEADER not in headers:
            raise SystemExit(f"Live xlsx missing column {setting.SUB_SKU_HEADER!r}")
        sub_ix = headers.index(setting.SUB_SKU_HEADER)
        st_ix = headers.index("scrape_status")

        by_sku: dict[str, dict[str, str]] = {}
        for r in range(2, ws.max_row + 1):
            status = ws.cell(row=r, column=st_ix + 1).value
            if str(status or "").strip().lower() != "ok":
                continue
            sub_cell = ws.cell(row=r, column=sub_ix + 1).value
            sku = str(sub_cell or "").strip()
            if not sku or "," in sku:
                continue
            scrape: dict[str, str] = {}
            for key in SCRAPED_COLUMN_KEYS:
                if key not in headers:
                    continue
                cix = headers.index(key) + 1
                val = ws.cell(row=r, column=cix).value
                scrape[key] = "" if val is None else str(val)
            by_sku[_norm_sku_match(sku)] = scrape
        return by_sku, headers, n_base
    finally:
        wb.close()


def _stamp_from_singles_path(singles_path: Path) -> str:
    for pat in (
        r"Homelagance-(\d{2}-\d{2}-\d{2}_\d{6})-single-subskus-sheets\.xlsx$",
        r"Homelagance-(\d{8}_\d{6})-single-subskus-sheets\.xlsx$",  # legacy YYYYMMDD_HHMMSS
    ):
        m = re.match(pat, singles_path.name, re.IGNORECASE)
        if m:
            return m.group(1)
    return datetime.now().strftime(setting.RUN_FILE_STAMP_FORMAT)


def _multiples_output_paths(stamp: str) -> tuple[Path, Path]:
    d = setting.LIVE_SUCCESS_XLSX_DIR
    d.mkdir(parents=True, exist_ok=True)
    from_singles = d / f"Homelagance-{stamp}-multiple-subskus-from-singles.xlsx"
    from_web = d / f"Homelagance-{stamp}-multiple-subskus-sheets.xlsx"
    return from_singles, from_web


def all_sub_skus_in_singles(
    sub_skus: list[str], singles_by_sku: dict[str, dict[str, str]]
) -> bool:
    if not sub_skus:
        return False
    for s in sub_skus:
        if _norm_sku_match(s) not in singles_by_sku:
            return False
    return True


def narrow_base_values(base_header: list[str], base_row: list[Any]) -> list[Any]:
    out: list[Any] = []
    for h in setting.MULTI_NARROW_BASE_HEADERS:
        if h in base_header:
            ix = base_header.index(h)
            out.append(base_row[ix] if ix < len(base_row) else None)
        else:
            out.append("")
    return out


def append_narrow_multi_row(
    ws: Any, excel_row: int, base_header: list[str], base_row: list[Any], attrs_json: str
) -> None:
    vals = narrow_base_values(base_header, base_row)
    for j, v in enumerate(vals, start=1):
        ws.cell(row=excel_row, column=j, value=v)
    ws.cell(row=excel_row, column=len(vals) + 1, value=attrs_json)


def init_narrow_multiples_workbook(out_path: Path) -> tuple[Any, Any]:
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb = openpyxl.Workbook()
    ws = wb.active
    assert ws is not None
    ws.title = "Sheet1"
    headers = list(setting.MULTI_NARROW_BASE_HEADERS) + ["attributes"]
    for j, h in enumerate(headers, start=1):
        ws.cell(row=1, column=j, value=h)
    wb.save(out_path)
    return wb, ws


def run(
    singles_xlsx_path: Path,
) -> tuple[Path | None, Path | None, int, int, int]:
    """
    Reads singles scrape output; writes up to two workbooks (same stamp as singles file).
    Returns (narrow_path, web_path, narrow_rows, web_rows_total, web_rows_ok).
    web_rows_ok is scrape_status ok on the web workbook; narrow has no status column (all count).
    """
    _load_env()
    login_url = os.environ.get("LOGIN_URL", "").strip()
    username = os.environ.get("USERNAME", "").strip()
    password = os.environ.get("PASSWORD", "").strip()
    if not login_url or not username or not password:
        raise SystemExit(
            "Set LOGIN_URL, USERNAME, and PASSWORD in HomoleganceTool/.env"
        )

    if not singles_xlsx_path.is_file():
        raise SystemExit(f"Singles xlsx not found: {singles_xlsx_path}")

    base_path = setting.get_base_sheet_path()
    if not base_path.is_file():
        raise SystemExit(f"Base Excel not found: {base_path}")

    base_header, data_rows = read_sheet_rows(base_path)
    if setting.SUB_SKU_HEADER not in base_header:
        raise SystemExit(f"Missing column {setting.SUB_SKU_HEADER!r} in base sheet")
    if setting.MASTER_SKU_HEADER not in base_header:
        raise SystemExit(f"Missing column {setting.MASTER_SKU_HEADER!r} in base sheet")

    sub_col = base_header.index(setting.SUB_SKU_HEADER)
    master_col = base_header.index(setting.MASTER_SKU_HEADER)

    multi_total = count_multi_sub_sku_rows(data_rows, sub_col)
    max_cap = setting.MAX_SKUS
    jobs_planned = multi_total if max_cap is None else min(multi_total, max_cap)

    stamp = _stamp_from_singles_path(singles_xlsx_path)
    path_narrow, path_web = _multiples_output_paths(stamp)

    log.info("[bold white on blue] ═══ Homelegance multi Sub-SKU (split outputs) ═══ [/]")
    log.info("[green]Singles source[/] → [cyan]%s[/cyan]", singles_xlsx_path)
    log.info(
        "[yellow]Multi Sub-SKU[/] rows (comma in Sub-SKU): [bold cyan]%d[/]",
        multi_total,
    )
    if max_cap is not None:
        log.info(
            "[yellow]MAX_SKUS[/]=[bold]%d[/] → up to [bold green]%d[/] multi rows",
            max_cap,
            jobs_planned,
        )
    log.info(
        "[dim]Narrow (all in singles)[/] → [magenta]%s[/]",
        path_narrow.name,
    )
    log.info("[dim]Full scrape (needs site)[/] → [magenta]%s[/]", path_web.name)

    singles_by_sku, _, n_base = load_singles_by_subsku(singles_xlsx_path)

    path_out_narrow: Path | None = None
    path_out_web: Path | None = None
    wb_narrow: Any = None
    ws_narrow: Any = None
    next_narrow = 2
    wb_web: Any = None
    ws_web: Any = None
    next_web = 2

    multi_attempts = 0

    if multi_total == 0:
        log.info("[dim]No multi Sub-SKU rows — nothing to write.[/]")
        return None, None, 0, 0, 0

    jobs_s: list[tuple[list[Any], list[str]]] = []
    jobs_w: list[tuple[list[Any], list[str], str, str | None]] = []

    for row in data_rows:
        sku_cell = row[sub_col] if sub_col < len(row) else None
        if is_single_sub_sku(sku_cell):
            continue
        if sku_cell is None or str(sku_cell).strip() == "":
            continue

        if max_cap is not None and multi_attempts >= max_cap:
            log.info(
                "[yellow]Stopped[/] — reached [bold]MAX_SKUS=%d[/] (remaining multi rows skipped)",
                max_cap,
            )
            break

        multi_attempts += 1
        sub_skus = parse_multi_sub_skus(sku_cell)
        master_raw = row[master_col] if master_col < len(row) else None
        search_term = normalize_master_sku_for_search(master_raw)
        master_display = str(master_raw or "")

        if not sub_skus:
            jobs_w.append((row, sub_skus, search_term, master_display))
            continue

        if all_sub_skus_in_singles(sub_skus, singles_by_sku):
            jobs_s.append((row, sub_skus))
        else:
            jobs_w.append((row, sub_skus, search_term, master_display))

    empty_scrape = {
        k: "" for k in SCRAPED_COLUMN_KEYS if k not in ("scrape_status", "scrape_error")
    }

    if jobs_s:
        path_out_narrow = path_narrow
        wb_narrow, ws_narrow = init_narrow_multiples_workbook(path_narrow)
        try:
            job_i = 0
            for row, sub_skus in jobs_s:
                job_i += 1
                attrs = build_attributes_json(sub_skus, [], singles_by_sku)
                append_narrow_multi_row(ws_narrow, next_narrow, base_header, row, attrs)
                wb_narrow.save(path_narrow)
                log.info(
                    "[bold green]Multi (singles-only)[/] [cyan]%d[/]/[white]%d[/] | Sub-SKUs [dim]%s[/] "
                    "[dim](no site — attributes from singles merge / model-only)[/]",
                    job_i,
                    len(jobs_s),
                    _esc(", ".join(sub_skus)),
                )
                log_attributes_breakdown(
                    sub_skus, [], singles_by_sku, mode="singles-only (no scrape)"
                )
                next_narrow += 1
        finally:
            wb_narrow.close()

    if not jobs_w:
        log.info(
            "[bold white on green] ═══ Multi pass done (no site rows) ═══ [/] "
            "rows written (narrow)=[cyan]%d[/]",
            len(jobs_s),
        )
        return path_out_narrow, path_out_web, len(jobs_s), 0, 0

    path_out_web = path_web
    web_rows_ok = 0
    wb_web, ws_web, _ = init_success_only_workbook(path_web, base_header)
    try:
        origin = _origin_from_login_url(login_url)
        playwright = sync_playwright().start()
        try:
            browser = playwright.chromium.launch(
                headless=setting.HEADLESS,
                slow_mo=setting.SLOW_MO_MS,
            )
            context = browser.new_context(
                viewport={"width": 1400, "height": 900},
                ignore_https_errors=True,
            )
            page = context.new_page()
            page.set_default_timeout(setting.DEFAULT_TIMEOUT_MS)

            login_once(page, login_url, username, password)

            job_i = 0
            for row, sub_skus, search_term, master_display in jobs_w:
                job_i += 1
                log.info(
                    "[bold magenta]Multi (web)[/] [cyan]%d[/]/[white]%d[/] | Master [yellow]%s[/] "
                    "| search [green]%s[/] | Sub-SKUs [dim]%s[/] "
                    "[dim](will scrape site if search succeeds)[/]",
                    job_i,
                    len(jobs_w),
                    _esc(master_display or ""),
                    _esc(search_term or "(empty)"),
                    _esc(", ".join(sub_skus)),
                )

                if not sub_skus:
                    err = "Sub-SKU cell has no tokens after split"
                    data = dict(empty_scrape)
                    data["attributes"] = json.dumps([], ensure_ascii=False)
                    append_scrape_row(ws_web, next_web, n_base, row, data, ok=False, error=err)
                    wb_web.save(path_web)
                    log.info(
                        "[yellow]attributes[/] [web (invalid Sub-SKU cell)] — empty list, not scraped"
                    )
                    next_web += 1
                    time.sleep(setting.DELAY_BETWEEN_SKUS_SEC)
                    continue

                if not search_term:
                    err = "NEW/Master SKU empty after normalize"
                    data = dict(empty_scrape)
                    data["attributes"] = build_attributes_json(sub_skus, [], singles_by_sku)
                    append_scrape_row(ws_web, next_web, n_base, row, data, ok=False, error=err)
                    wb_web.save(path_web)
                    log_attributes_breakdown(
                        sub_skus, [], singles_by_sku, mode="web (skipped — empty master)"
                    )
                    next_web += 1
                    time.sleep(setting.DELAY_BETWEEN_SKUS_SEC)
                    continue

                try:
                    log.info("[blue]Search[/] → master SKU → [dim]submit[/]")
                    search_sku(page, search_term)
                    log.info("[blue]Search[/] → [green]results[/] — exact card…")
                    if click_exact_product_card(page, search_term):
                        log.info(
                            "[bold magenta]SCRAPING[/] PDP + packaging table in browser "
                            "(product fields + Element grid)…"
                        )
                        pdata = scrape_product_page(page, origin)
                        table = scrape_packaging_el_table(page)
                        pdata["attributes"] = build_attributes_json(
                            sub_skus, table, singles_by_sku
                        )
                        append_scrape_row(ws_web, next_web, n_base, row, pdata, ok=True)
                        web_rows_ok += 1
                        log_attributes_breakdown(
                            sub_skus, table, singles_by_sku, mode="web scrape (ok)"
                        )
                        log.info(
                            "[bold green]XLSX[/] row [cyan]%d[/] | scraped packaging rows=[white]%d[/]",
                            next_web,
                            len(table),
                        )
                        next_web += 1
                        wb_web.save(path_web)
                    else:
                        err = "No exact product card for master search"
                        data = dict(empty_scrape)
                        data["attributes"] = build_attributes_json(
                            sub_skus, [], singles_by_sku
                        )
                        append_scrape_row(ws_web, next_web, n_base, row, data, ok=False, error=err)
                        next_web += 1
                        wb_web.save(path_web)
                        log.warning("[red]%s[/] — [yellow]%s[/]", err, _esc(search_term))
                        log_attributes_breakdown(
                            sub_skus, [], singles_by_sku, mode="web (no PDP — not scraped)"
                        )
                except Exception as e:  # noqa: BLE001
                    err = str(e)
                    data = dict(empty_scrape)
                    data["attributes"] = build_attributes_json(sub_skus, [], singles_by_sku)
                    append_scrape_row(ws_web, next_web, n_base, row, data, ok=False, error=err)
                    next_web += 1
                    wb_web.save(path_web)
                    log.warning(
                        "[bold red]ERROR[/] — Master [yellow]%s[/]: [red]%s[/]",
                        _esc(search_term),
                        _esc(err),
                    )
                    log_attributes_breakdown(
                        sub_skus, [], singles_by_sku, mode="web (error — not scraped)"
                    )
                    try:
                        page.goto(origin + "/", wait_until="domcontentloaded")
                    except Exception:
                        pass

                time.sleep(setting.DELAY_BETWEEN_SKUS_SEC)

            context.close()
            browser.close()
        finally:
            playwright.stop()
    finally:
        wb_web.close()

    web_total = len(jobs_w)
    narrow_n = len(jobs_s)
    log.info(
        "[bold white on green] ═══ Multi scraper finished ═══ [/] "
        "narrow rows=[cyan]%d[/] | web rows=[cyan]%d[/] ([green]ok[/]=[bold]%d[/] "
        "[red]failed[/]=[bold]%d[/])",
        narrow_n,
        web_total,
        web_rows_ok,
        web_total - web_rows_ok,
    )
    return path_out_narrow, path_out_web, narrow_n, web_total, web_rows_ok


if __name__ == "__main__":
    from log_theme import setup_colored_logging

    setup_colored_logging()
    raise SystemExit(
        "Run via run.py after singles, or: "
        "from multiplesNewMaterskuNaserSubskuFindings import run; "
        "run(Path('...single-subskus-sheets.xlsx'))"
    )
